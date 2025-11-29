import streamlit as st
import pandas as pd
import openpyxl
from PIL import Image
import base64
import os
import io

st.set_page_config(
    page_title="Report Card System",
    page_icon="🏫",
    layout="wide"
)

# Load CSS for styling
st.markdown("""
<style>
.main-header {
    text-align: center;
    color: #2E86AB;
    font-size: 2.5rem;
    margin-bottom: 2rem;
}
.student-card {
    background: #f0f2f6;
    padding: 1rem;
    border-radius: 10px;
    margin: 1rem 0;
}
</style>
""", unsafe_allow_html=True)

def load_students_from_files(mid_file, final_file):
    """Load student data from uploaded Excel files"""
    try:
        students = {}
        
        # Load workbooks from uploaded files
        mid_wb = openpyxl.load_workbook(io.BytesIO(mid_file.read()))
        final_wb = openpyxl.load_workbook(io.BytesIO(final_file.read()))
        
        # Get common sheets
        common_sheets = set(mid_wb.sheetnames) & set(final_wb.sheetnames)
        
        for sheet_name in common_sheets:
            mid_sheet = mid_wb[sheet_name]
            final_sheet = final_wb[sheet_name]
            
            # Find header row (assuming row 1 has headers)
            for row in range(1, min(10, mid_sheet.max_row + 1)):
                cell_value = str(mid_sheet.cell(row, 1).value or "").strip().lower()
                if any(keyword in cell_value for keyword in ['roll', 'student', 'name']):
                    header_row = row
                    break
            else:
                header_row = 1
            
            # Extract student data
            for row in range(header_row + 1, mid_sheet.max_row + 1):
                roll_no = mid_sheet.cell(row, 1).value
                name = mid_sheet.cell(row, 2).value
                
                if roll_no and name:
                    student_key = f"{sheet_name}_{roll_no}"
                    students[student_key] = {
                        'name': str(name),
                        'roll_no': str(roll_no),
                        'class': sheet_name,
                        'mid_marks': {},
                        'final_marks': {}
                    }
                    
                    # Get subject marks from both sheets
                    for col in range(3, mid_sheet.max_column + 1):
                        subject = mid_sheet.cell(header_row, col).value
                        if subject:
                            mid_mark = mid_sheet.cell(row, col).value
                            final_mark = final_sheet.cell(row, col).value
                            
                            students[student_key]['mid_marks'][str(subject)] = mid_mark or 0
                            students[student_key]['final_marks'][str(subject)] = final_mark or 0
        
        return students
    except Exception as e:
        st.error(f"Error loading files: {str(e)}")
        return {}

def generate_report_card(student_data):
    """Generate HTML report card for a student"""
    html = f"""
    <div style="border: 2px solid #2E86AB; padding: 20px; margin: 10px; border-radius: 10px;">
        <h2 style="text-align: center; color: #2E86AB;">Report Card</h2>
        <p><strong>Name:</strong> {student_data['name']}</p>
        <p><strong>Roll No:</strong> {student_data['roll_no']}</p>
        <p><strong>Class:</strong> {student_data['class']}</p>
        
        <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
            <tr style="background-color: #2E86AB; color: white;">
                <th style="border: 1px solid #ddd; padding: 8px;">Subject</th>
                <th style="border: 1px solid #ddd; padding: 8px;">Mid Term</th>
                <th style="border: 1px solid #ddd; padding: 8px;">Final Term</th>
                <th style="border: 1px solid #ddd; padding: 8px;">Total</th>
            </tr>
    """
    
    total_marks = 0
    subject_count = 0
    
    for subject in student_data['mid_marks']:
        mid = student_data['mid_marks'].get(subject, 0)
        final = student_data['final_marks'].get(subject, 0)
        total = mid + final
        total_marks += total
        subject_count += 1
        
        html += f"""
            <tr>
                <td style="border: 1px solid #ddd; padding: 8px;">{subject}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{mid}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{final}</td>
                <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{total}</td>
            </tr>
        """
    
    percentage = (total_marks / (subject_count * 100)) * 100 if subject_count > 0 else 0
    grade = "A+" if percentage >= 90 else "A" if percentage >= 80 else "B" if percentage >= 70 else "C" if percentage >= 60 else "F"
    
    html += f"""
        </table>
        <p><strong>Total Marks:</strong> {total_marks}</p>
        <p><strong>Percentage:</strong> {percentage:.2f}%</p>
        <p><strong>Grade:</strong> {grade}</p>
    </div>
    """
    
    return html

# Main App
st.markdown('<h1 class="main-header">🏫 Report Card System</h1>', unsafe_allow_html=True)

# File upload section
st.header("Upload Excel Files")
col1, col2 = st.columns(2)

with col1:
    mid_file = st.file_uploader("Upload Mid Term Excel File", type=['xlsx', 'xls'])

with col2:
    final_file = st.file_uploader("Upload Final Term Excel File", type=['xlsx', 'xls'])

if mid_file and final_file:
    with st.spinner("Loading student data..."):
        students = load_students_from_files(mid_file, final_file)
    
    if students:
        st.success(f"Loaded {len(students)} students successfully!")
        
        # Student selection
        st.header("Generate Report Card")
        student_options = {f"{data['name']} ({data['roll_no']}) - {data['class']}": key 
                          for key, data in students.items()}
        
        selected_student = st.selectbox("Select Student", list(student_options.keys()))
        
        if selected_student:
            student_key = student_options[selected_student]
            student_data = students[student_key]
            
            # Generate and display report card
            report_html = generate_report_card(student_data)
            st.markdown(report_html, unsafe_allow_html=True)
            
            # Download button
            st.download_button(
                label="Download Report Card (HTML)",
                data=report_html,
                file_name=f"report_card_{student_data['roll_no']}.html",
                mime="text/html"
            )
    else:
        st.error("No student data found. Please check your Excel files.")
else:
    st.info("Please upload both Mid Term and Final Term Excel files to get started.")
    
    # Show sample format
    st.header("Expected Excel Format")
    st.write("Your Excel files should have the following structure:")
    
    sample_data = {
        'Roll No': [1, 2, 3],
        'Student Name': ['John Doe', 'Jane Smith', 'Bob Johnson'],
        'Math': [85, 92, 78],
        'English': [88, 85, 82],
        'Science': [90, 88, 85]
    }
    
    st.dataframe(pd.DataFrame(sample_data))
